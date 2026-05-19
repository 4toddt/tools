#!/usr/bin/env python3
"""Build per-SKU master cover CSV from MASTER FABRICS and MASTER LEATHER tabs.

Generates ``data/cover-skus-by-color.csv`` with one row per cover SKU id (e.g. B153808),
fully decoded against the workbook's legends, with dropStatus recovered from
per-color-line strikethrough or red font (FF0000) inside the Colors cell's
rich-text runs.

Run from the project folder (venv lives next to ``scripts/``):

    cd docs/projects/bcc-cover-sku-reconciliation
    .xlsx-venv/bin/python scripts/build-by-sku-csv.py
"""
import csv
import re
from collections import Counter
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock

PROJECT = Path(__file__).resolve().parent.parent
WORKBOOK = PROJECT / "data" / "Copy of April 2026 Cover Information List FINAL (v.4.10.2026).xlsx"
OUTPUT = PROJECT / "data" / "cover-skus-by-color.csv"


WEARABILITY = {"N": "Normal", "M": "Medium", "P": "Performance"}

SPECIALTY_FABRIC = {
    "I": "iClean",
    "C": "Conserve",
    "A": "Antimicrobial",
    "PF": "Pet Friendly",
    "N": "Nanobionic",
    "W*B": "Bleach Cleanable",
}

LEATHER_DOT = {"A": "Authentic", "P": "Performance", "N": "Nubuck"}

LEATHER_PRODUCT_LINE = {
    "SLR": "Select Leather Reclining",
    "SLS": "Select Leather Stationary",
    "CLM": "Custom Leather Match",
    "CLP": "Custom Leather Plus",
    "SL": "Specialty Leather",
}

SPECIALTY_LEATHER = {"A": "Antimicrobial", "C": "Crypton", "I": "iClean"}

COLUMNS = [
    "skuId",
    "coverType",
    "seriesNumber",
    "patternName",
    "colorCode",
    "colorName",
    "productLineCode",
    "productLineDescription",
    "introDate",
    "cleaningCode",
    "description",
    "capCode",
    "faceContents",
    "backContents",
    "wearabilityCode",
    "wearabilityDescription",
    "specialtyFabricCode",
    "specialtyFabricDescription",
    "weltPattern",
    "slipCoverProgram",
    "horizRepeat",
    "vertRepeat",
    "leatherDotCode",
    "leatherDotDescription",
    "specialtyLeatherCode",
    "specialtyLeatherDescription",
    "marriedStyles",
    "notes",
    "postageStampSize",
    "dropStatus",
    "issues",
    "sourceSheet",
    "sourceRow",
]

COLOR_LINE_RE = re.compile(r"^(\d{2})\s*([=\-])\s*(.+?)\s*$")
SERIES_RE = re.compile(r"^[A-Z]+\d+$")


def _scalar_text(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def cell_text(cell):
    """Return the cell text, joining rich-text runs without losing newlines."""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, CellRichText):
        return "".join(str(p) for p in v)
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v)


def per_line_format(cell):
    """Walk rich-text runs and return a list (one entry per newline-separated line).

    Each entry is a list of (char, dropped_bool) tuples where dropped is True if the
    char's run has strikethrough or red (FF0000) font color. Returns [] for plain
    text cells (no rich text to interpret).
    """
    v = cell.value
    if not isinstance(v, CellRichText):
        return []

    chars = []
    for part in v:
        if isinstance(part, TextBlock):
            font = part.font
            dropped = False
            if font:
                if getattr(font, "strike", False):
                    dropped = True
                color = getattr(font, "color", None)
                if color is not None:
                    rgb = getattr(color, "rgb", None)
                    if isinstance(rgb, str) and rgb.upper().endswith("FF0000"):
                        dropped = True
            for ch in str(part):
                chars.append((ch, dropped))
        else:
            for ch in str(part):
                chars.append((ch, False))

    lines = []
    current = []
    for ch, dropped in chars:
        if ch == "\n":
            lines.append(current)
            current = []
        else:
            current.append((ch, dropped))
    lines.append(current)
    return lines


def classify_line_drop(line_chars):
    """Return ('dropped'|'', mixed_bool) for a single line's char/dropped tuples."""
    non_ws = [d for c, d in line_chars if not c.isspace()]
    if not non_ws:
        return "", False
    if all(non_ws):
        return "dropped", False
    if any(non_ws):
        return "dropped", True
    return "", False


def parse_color_block(cell_str):
    """Yield (line_index, code, name, issue) for each non-empty line in the block.

    line_index is the index when the cell text is split on '\n', so the caller can
    correlate parsed colors with per-line formatting from per_line_format().
    """
    if not cell_str:
        return
    for idx, line in enumerate(cell_str.split("\n")):
        stripped = line.strip()
        if not stripped:
            continue
        m = COLOR_LINE_RE.match(stripped)
        if not m:
            yield idx, None, None, "unparsed-color-line"
            continue
        code = m.group(1)
        sep = m.group(2)
        name = re.sub(r"\s+", " ", m.group(3).strip())
        issue = "hyphen-separator" if sep == "-" else None
        yield idx, code, name, issue


def decode_specialty_fabric(raw):
    if raw is None or str(raw).strip() == "":
        return "", ""
    parts = [p.strip() for p in str(raw).split(",") if p.strip()]
    decoded = [SPECIALTY_FABRIC.get(p, p) for p in parts]
    return ", ".join(parts), ", ".join(decoded)


def decode_leather_product_line(raw):
    if raw is None or str(raw).strip() == "":
        return "", ""
    pieces = [p.strip() for p in re.split(r"[,\n]", str(raw)) if p.strip()]
    decoded = [LEATHER_PRODUCT_LINE.get(p, p) for p in pieces]
    return "; ".join(pieces), "; ".join(decoded)


def decode_specialty_leather(raw):
    if raw is None or str(raw).strip() == "":
        return "", ""
    pieces = [p.strip() for p in re.split(r"[,/\n]", str(raw)) if p.strip()]
    decoded = [SPECIALTY_LEATHER.get(p, p) for p in pieces]
    return ", ".join(pieces), ", ".join(decoded)


def collect_series_counts(ws, header_row, series_col):
    counts = Counter()
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(r, series_col).value
        if not v:
            continue
        s = str(v).strip()
        if SERIES_RE.match(s):
            counts[s] += 1
    return counts


def build_fabric_rows(ws):
    rows = []
    series_counts = collect_series_counts(ws, header_row=12, series_col=1)

    for r in range(13, ws.max_row + 1):
        series_raw = ws.cell(r, 1).value
        if not series_raw:
            continue
        series = str(series_raw).strip()
        if not SERIES_RE.match(series):
            continue

        pattern_name = _scalar_text(ws.cell(r, 2).value)
        colors_cell = ws.cell(r, 3)
        product_line = _scalar_text(ws.cell(r, 5).value)
        intro = _scalar_text(ws.cell(r, 6).value)
        horiz = _scalar_text(ws.cell(r, 7).value)
        vert = _scalar_text(ws.cell(r, 8).value)
        cap = _scalar_text(ws.cell(r, 9).value)
        face = _scalar_text(ws.cell(r, 10).value)
        back = _scalar_text(ws.cell(r, 11).value)
        wear_raw = _scalar_text(ws.cell(r, 12).value)
        spec_fab_raw = ws.cell(r, 13).value
        welt = _scalar_text(ws.cell(r, 14).value)
        slip = _scalar_text(ws.cell(r, 15).value)
        clean = _scalar_text(ws.cell(r, 16).value)
        desc = _scalar_text(ws.cell(r, 17).value)

        line_format = per_line_format(colors_cell)
        cell_str = cell_text(colors_cell)
        spec_code, spec_desc = decode_specialty_fabric(spec_fab_raw)

        for idx, code, name, parse_issue in parse_color_block(cell_str):
            issues = []
            if parse_issue:
                issues.append(parse_issue)

            if 0 <= idx < len(line_format):
                drop_status, mixed = classify_line_drop(line_format[idx])
                if mixed:
                    issues.append("formatting-ambiguous")
            else:
                drop_status = ""
                if line_format:
                    issues.append("formatting-ambiguous")

            if series_counts[series] > 1:
                issues.append("duplicate-series")

            sku_id = f"{series}{code}" if code else ""
            rows.append(
                {
                    "skuId": sku_id,
                    "coverType": "fabric",
                    "seriesNumber": series,
                    "patternName": pattern_name,
                    "colorCode": code or "",
                    "colorName": name or "",
                    "productLineCode": product_line,
                    "productLineDescription": "",
                    "introDate": intro,
                    "cleaningCode": clean,
                    "description": desc,
                    "capCode": cap,
                    "faceContents": face,
                    "backContents": back,
                    "wearabilityCode": wear_raw,
                    "wearabilityDescription": WEARABILITY.get(wear_raw, ""),
                    "specialtyFabricCode": spec_code,
                    "specialtyFabricDescription": spec_desc,
                    "weltPattern": welt,
                    "slipCoverProgram": slip,
                    "horizRepeat": horiz,
                    "vertRepeat": vert,
                    "leatherDotCode": "",
                    "leatherDotDescription": "",
                    "specialtyLeatherCode": "",
                    "specialtyLeatherDescription": "",
                    "marriedStyles": "",
                    "notes": "",
                    "postageStampSize": "",
                    "dropStatus": drop_status,
                    "issues": "|".join(issues),
                    "sourceSheet": "MASTER FABRICS",
                    "sourceRow": r,
                }
            )

    return rows


def build_leather_rows(ws):
    rows = []
    series_counts = collect_series_counts(ws, header_row=17, series_col=1)

    for r in range(18, ws.max_row + 1):
        series_raw = ws.cell(r, 1).value
        if not series_raw:
            continue
        series = str(series_raw).strip()
        if not SERIES_RE.match(series):
            continue

        pattern_name = _scalar_text(ws.cell(r, 2).value)
        product_line_raw = ws.cell(r, 3).value
        intro = _scalar_text(ws.cell(r, 4).value)
        colors_cell = ws.cell(r, 5)
        ps = _scalar_text(ws.cell(r, 6).value)
        leather_dot = _scalar_text(ws.cell(r, 8).value)
        spec_leather_raw = ws.cell(r, 10).value
        clean = _scalar_text(ws.cell(r, 11).value)
        married = cell_text(ws.cell(r, 12)).strip()
        notes = cell_text(ws.cell(r, 15)).strip()
        desc = _scalar_text(ws.cell(r, 18).value)

        line_format = per_line_format(colors_cell)
        cell_str = cell_text(colors_cell)
        pl_code, pl_desc = decode_leather_product_line(product_line_raw)
        sl_code, sl_desc = decode_specialty_leather(spec_leather_raw)

        for idx, code, name, parse_issue in parse_color_block(cell_str):
            issues = []
            if parse_issue:
                issues.append(parse_issue)

            if 0 <= idx < len(line_format):
                drop_status, mixed = classify_line_drop(line_format[idx])
                if mixed:
                    issues.append("formatting-ambiguous")
            else:
                drop_status = ""
                if line_format:
                    issues.append("formatting-ambiguous")

            if series_counts[series] > 1:
                issues.append("duplicate-series")

            sku_id = f"{series}{code}" if code else ""
            rows.append(
                {
                    "skuId": sku_id,
                    "coverType": "leather",
                    "seriesNumber": series,
                    "patternName": pattern_name,
                    "colorCode": code or "",
                    "colorName": name or "",
                    "productLineCode": pl_code,
                    "productLineDescription": pl_desc,
                    "introDate": intro,
                    "cleaningCode": clean,
                    "description": desc,
                    "capCode": "",
                    "faceContents": "",
                    "backContents": "",
                    "wearabilityCode": "",
                    "wearabilityDescription": "",
                    "specialtyFabricCode": "",
                    "specialtyFabricDescription": "",
                    "weltPattern": "",
                    "slipCoverProgram": "",
                    "horizRepeat": "",
                    "vertRepeat": "",
                    "leatherDotCode": leather_dot,
                    "leatherDotDescription": LEATHER_DOT.get(leather_dot, ""),
                    "specialtyLeatherCode": sl_code,
                    "specialtyLeatherDescription": sl_desc,
                    "marriedStyles": married,
                    "notes": notes,
                    "postageStampSize": ps,
                    "dropStatus": drop_status,
                    "issues": "|".join(issues),
                    "sourceSheet": "MASTER LEATHER",
                    "sourceRow": r,
                }
            )

    return rows


def main():
    wb = openpyxl.load_workbook(WORKBOOK, rich_text=True, data_only=True)
    rows = build_fabric_rows(wb["MASTER FABRICS"]) + build_leather_rows(wb["MASTER LEATHER"])
    wb.close()

    rows.sort(key=lambda x: (x["coverType"], x["seriesNumber"], x["colorCode"]))

    with OUTPUT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    n = len(rows)
    n_dropped = sum(1 for r in rows if r["dropStatus"] == "dropped")
    n_with_issues = sum(1 for r in rows if r["issues"])
    issue_counter = Counter()
    for r in rows:
        for tag in r["issues"].split("|"):
            if tag:
                issue_counter[tag] += 1

    print(f"Wrote {OUTPUT.name}: {n} rows, {n_dropped} dropped, {n_with_issues} with issues")
    if issue_counter:
        print("Issue breakdown:")
        for tag, count in issue_counter.most_common():
            print(f"  {tag}: {count}")
    print()

    spot = next((r for r in rows if r["skuId"] == "B153808"), None)
    print("Spot-check B153808:")
    if spot:
        for k in (
            "skuId",
            "coverType",
            "seriesNumber",
            "patternName",
            "colorCode",
            "colorName",
            "wearabilityCode",
            "wearabilityDescription",
            "capCode",
            "introDate",
            "dropStatus",
            "issues",
        ):
            print(f"  {k}: {spot[k]!r}")
    else:
        print("  NOT FOUND")
    print()

    dropped_examples = [r for r in rows if r["dropStatus"] == "dropped"]
    print(f"Dropped rows ({len(dropped_examples)}):")
    for r in dropped_examples[:20]:
        print(
            f"  {r['skuId']:<12} {r['coverType']:<7} "
            f"{r['patternName']:<20} {r['colorName']:<20} "
            f"row {r['sourceRow']} (issues={r['issues']!r})"
        )


if __name__ == "__main__":
    main()
